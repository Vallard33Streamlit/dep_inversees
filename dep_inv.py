import streamlit as st
import streamlit_antd_components as sac
import pandas as pd
import numpy as np
from copy import deepcopy
from io import BytesIO
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment


version_baci = "2026"
annee_baci = "2024"
version_hs6 = "2022"
version_aipnet = "04-Dec-2024"
nomenclature_aipnet = "HS6 2002"

def get_explication_filtre(filtre, type, type2):
    if filtre == "hhi_c" :
        return f"L'indice HHI des {type}ortations mesure la concentration des {type}ortations d'un pays entre ses partenaires. Un indice élevé indique une forte dépendance à un petit nombre de pays {type2}ortateurs, tandis qu'un indice faible traduit une plus grande diversification."
    elif filtre == "hhi_M" :
        return f"L'indice HHI des {type2}ortations mesure la concentration des {type2}ortations dans le monde. Un indice élevé indique une forte dépendance mondiale à un petit nombre de pays {type2}ortateurs, tandis qu'un indice faible traduit une plus grande diversification."
    elif filtre == "p_fr_ue_in_c" :
        return f"La part des {type2}ortations de {st.session_state.fr_ue_lab} dans les {type}ortations du pays mesure l'impact direct qu'une restriction de {st.session_state.fr_ue_lab} sur ses {type2}ortations aura sur les {type}ortations du pays étudié."
    elif filtre == "p_c_in_fr_ue" :
        return f"La part des {type}ortations du pays dans les {type2}ortations de {st.session_state.fr_ue_lab} mesure l'impact direct qu'une restriction de {st.session_state.fr_ue_lab} sur ses {type2}ortations envers le pays étudié aura sur ses propres {type2}ortations."
    elif filtre == "igpc_rank":
        return "Le score IGPC représente la centralité du produit dans les chaînes de valeurs mondiales. Plus le rang est proche de 100 plus le produit est central, plus il est proche de 0 moins il est central. Cette centralité est mesurée dans le papier AIPNET grâce à une reconstitution d'un graphe des chaînes de valeur par IA et grâce à la part du marché mondiale et la concentration du marché de chaque produit."
    else :
        return ""
    
st.set_page_config(layout="wide")
# --- 1. Initialisation des données (cachée) ---
st.header("Dépendances inversées")
def load_baci():
    l_bacis = []
    for i in range (5):
        l_bacis.append(pd.read_csv(f"baci{i+1}.csv", dtype={"k": "int32", "i": "int16", "j": "int16", "v": "float64", "q": "float64"}, usecols=["i", "j", "k", "v", "q"]))
    return pd.concat(l_bacis, ignore_index=True)

def conv_str(x):
    s = str(x)
    if len(s) == 5:
        return "0" + s
    else :
        return s

@st.cache_data
def load_countries_config():
    countries_config = pd.read_csv("country_codes.csv", sep=";")[["country_code", "nom_pays"]]

    countries_config["country"] = True
    countries_config["zone"] = np.nan
    l_ue = [276, 40, 56, 100, 196, 191, 208, 724, 233, 246, 251, 300, 348, 372, 380, 428, 440, 442, 470, 528, 616, 620, 642, 703, 705, 752, 203]
    countries_config.loc[1000] = 1000, "UE", False, np.nan
    countries_config.loc[countries_config['country_code'].isin(set(l_ue)), "zone"] = 1000
    return countries_config

@st.cache_data
def load_labels():
    labels_sections = pd.read_csv("labels_sections.csv")
    labels_sections["l_hs2"] = labels_sections["l_hs2"].fillna("").apply(lambda x : x.split(","))
    labels_sections.loc[labels_sections["Niveau"] == "HS2", "l_hs2"] = np.nan
    labels_sections_tree = []
    for s in labels_sections[labels_sections["Niveau"] == "Section"].iterrows():
        dic={}
        dic["label"] = "Section " + s[1].Catégorie + " - " + s[1].Label
        dic["value"] = s[1].Catégorie
        children = []
        for c in labels_sections[labels_sections["Catégorie"].apply(lambda x : x in s[1].l_hs2)].iterrows():
            dic_c = {}
            dic_c["label"] = c[1].Catégorie + " - " + c[1].Label
            dic_c["value"] = c[1].Catégorie
            children.append(dic_c)
        dic["children"] = children
        labels_sections_tree.append(dic)
    return pd.read_csv("labels_hs6.csv", dtype={"Code HS6": str}), labels_sections, labels_sections_tree, list(labels_sections.loc[(labels_sections["Niveau"] == "HS2")&(labels_sections["Catégorie"] >= "25"), "Catégorie"].index)

@st.cache_data
def load_aipnet_conv02():
    aipnet = pd.read_excel("AIPNET_Data_Pack_20241204.xlsx", sheet_name="1a. Node List 6-digit HS02")[["hs_code_6d", "IGPC", "IGPC_rank"]]
    conv02 = pd.read_excel("HS2022toHS2002ConversionAndCorrelationTables.xlsx", sheet_name="HS2022-HS2002 Conversions")[["From HS 2022", "From HS 2002"]]
    return aipnet, conv02
            

def calc_baci2(countries_config):
    baci2 = load_baci().merge(countries_config.loc[countries_config["zone"].notna(), ["country_code", "zone"]], left_on="i", right_on="country_code", how="left")
    baci2.loc[baci2["zone"].notna(), "i"] = baci2.loc[baci2["zone"].notna(), "zone"] 
    baci2.drop(columns=["country_code", "zone"], inplace=True)
    baci2 = baci2.merge(countries_config.loc[countries_config["zone"].notna(), ["country_code", "zone"]], left_on="j", right_on="country_code", how="left")
    baci2.loc[baci2["zone"].notna(), "j"] = baci2.loc[baci2["zone"].notna(), "zone"]
    baci2.drop(columns=["country_code", "zone"], inplace=True)
    baci2 = baci2[baci2["i"] != baci2['j']]
    return baci2.groupby(["i","j","k"])[["v", "q"]].sum().reset_index()

def calc_var(type, country=0):
    if type == "imp":
        x, y = "j", "i"
        type_o = "exp"
    else:
        x, y = "i", "j"
        type_o = "imp"
    if country == 0:
        type2="m"
        baci3 = st.session_state.baci2.groupby(["k", y])[["v", "q"]].sum().reset_index()
    else:
        type2 = "c"
        baci3 = st.session_state.baci2[st.session_state.baci2[x] == country].copy()

    hhi_c =baci3.groupby("k").agg(
            v=("v", "sum"),
            q=("q", "sum"),
            hhi=("v", lambda x: (x ** 2).sum())
    ).rename(columns={col : f"{col}_{type}_{type2}" for col in ["v", "q", "hhi"]}).reset_index()
    
    temp_fr_ue = baci3.loc[baci3[y] == st.session_state.fr_ue, ["k", "v"]].groupby("k").sum().reset_index().rename(columns={"v":f"p_{type_o}_fr_ue_{type}_{type2}"})

    temp = baci3[["k", y, "v"]].sort_values("v", ascending=False).groupby("k").head(3).set_index("k")
    del baci3
    temp["n"] = temp.groupby(level="k").cumcount() + 1
    temp = temp.reset_index(level="k")
    top3 = temp.pivot(index="k", columns="n")
    top3.columns = [f"{col}{n}" for col, n in top3.columns]
    top3.reset_index(inplace=True)
    for col in top3.columns:
        if col[0]==y:
            top3 = top3.merge(c_config, left_on=col, right_on="country_code", how="left").rename(columns={"nom_pays":f"c_{type_o}_max_{col[-1]}_{type2}_{type}"})
            top3.drop(columns=[col, "country_code", "country", "zone"], inplace=True)
    top3.rename(columns={col : f"cc_{type_o}_max_{col[-1]}_{type2}_{type}" for col in top3.columns if col[0]==y}, inplace=True)
    top3.rename(columns={col : f"p_{type_o}_max_{col[-1]}_{type2}_{type}" for col in top3.columns if col[0]=="v"}, inplace=True)

    df_tot = hhi_c.merge(top3, how="outer")
    del hhi_c, top3
    df_tot[f"hhi_{type}_{type2}"] /= (df_tot[f"v_{type}_{type2}"]**2)
    for col in df_tot.columns:
        if col[0] == "p":
            df_tot[col] = df_tot[col].fillna(0) / df_tot[f"v_{type}_{type2}"] * 100

    df_tot = df_tot.merge(temp_fr_ue, how="outer")
    df_tot[f"p_{type_o}_fr_ue_{type}_{type2}"] = df_tot[f"p_{type_o}_fr_ue_{type}_{type2}"].fillna(0) / df_tot[f"v_{type}_{type2}"] * 100
    
    if country != 0:
        baci4 = st.session_state.baci2[st.session_state.baci2[y] == st.session_state.fr_ue]
        temp = baci4[baci4[x] == country].merge(baci4.groupby("k")["v"].sum().rename("v_sum").reset_index(), how="outer")
        temp[f"p_{type}_{type2}_{type_o}_fr_ue"] = temp["v"].fillna(0) / temp["v_sum"] * 100
        df_tot = df_tot.merge(temp[["k", f"p_{type}_{type2}_{type_o}_fr_ue"]], how="outer")
    return df_tot

def maj(c):
    return c[0].upper() + c[1:]

if "compt_z_infl" not in st.session_state:
    st.session_state.compt_z_infl = 2

countries_config = load_countries_config()
labels, labels_sections, labels_sections_tree, checked_sections = load_labels()

if "countries_config" not in st.session_state:
    st.session_state.countries_config = deepcopy(countries_config)
c_config = st.session_state.countries_config
if "modified_z_infl" not in st.session_state:
    st.session_state.modified_z_infl = True  # Flag pour savoir si des modifications sont en attente
if "modified_sel_country" not in st.session_state:
    st.session_state.modified_sel_country = True


with st.expander("**Éditer les zones d'influence**", expanded=True):
    # --- Ajout d'une nouvelle zone ---
    new_zone_name = st.text_input("Nom de la nouvelle zone", value ="Zone d'influence", key=f"new_zone_input")
    if st.button("➕ Ajouter une zone", key="add_zone_btn") and new_zone_name:
        if new_zone_name == "Zone d'influence":
            new_zone_name += f" {st.session_state.compt_z_infl}"
        if new_zone_name not in c_config["nom_pays"].values:
            c_config.loc[1000+st.session_state.compt_z_infl] = 1000+st.session_state.compt_z_infl, new_zone_name, False, np.nan
            st.session_state.modified_z_infl = True
            st.session_state.compt_z_infl += 1
        else:
            st.error(new_zone_name + " existe déjà.")
    # --- Édition des zones existantes ---
    for zone in list(c_config.loc[~c_config["country"], "country_code"].unique()):
        col1, col2, col3 = st.columns([1, 8, 1])
        with col1:
            # Renommage de la zone
            new_name = st.text_input(
                "Nom de la zone",
                value=c_config.loc[c_config["country_code"] == zone, "nom_pays"].iloc[0],
                key=f"zone_name_{zone}",
                label_visibility="collapsed"
            )
            if new_name != c_config.loc[c_config["country_code"] == zone, "nom_pays"].iloc[0]:
                if new_name not in c_config["nom_pays"].values:
                    c_config.loc[c_config["country_code"] == zone, "nom_pays"] = new_name
                    st.session_state.modified_z_infl = True
                else:
                    st.error(new_name + " existe déjà.")
        with col2:
            # Sélection des pays pour cette zone
            selected_countries = st.multiselect(
                "Pays associés",
                options=sorted(list(c_config.loc[(c_config["country"]&c_config["zone"].isna()) | (c_config["zone"] == zone), "nom_pays"])),
                default=sorted(list(c_config.loc[c_config["zone"] == zone, "nom_pays"])),
                key=f"countries_{zone}",
                label_visibility="collapsed"
            )
            temp = set(c_config.loc[c_config["zone"] == zone, "nom_pays"])
            if set(selected_countries) != temp:
                c_config.loc[c_config['nom_pays'].isin(set(selected_countries) - temp),"zone"] = zone
                c_config.loc[c_config['nom_pays'].isin(temp - set(selected_countries)), "zone"] = np.nan
                st.session_state.modified_z_infl = True

        with col3:
            # Suppression de la zone
            if st.button("🗑️", key=f"del_zone_{zone}"):
                c_config.drop(zone, inplace=True)
                st.session_state.modified_z_infl = True
                st.rerun()

    # --- Bouton de validation globale ---
    if st.session_state.modified_z_infl:
        if st.button("💾 **Enregistrer la configuration**", type="primary"):
            st.session_state.modified_sel_country = True
            if c_config.loc[c_config["country_code"] == 251, "zone"].isna().any():
                st.session_state.fr_ue = 251
            else :
                st.session_state.fr_ue = c_config.loc[c_config["country_code"] == 251, "zone"].iloc[0]
            st.session_state.baci2 = calc_baci2(c_config)
            st.session_state.df_m = calc_var("imp", 0).merge(calc_var("exp", 0), how="outer")
            aipnet, conv02 = load_aipnet_conv02()
            st.session_state.df_m = st.session_state.df_m.merge(conv02, how="left", left_on="k", right_on="From HS 2022")
            st.session_state.df_m.loc[st.session_state.df_m["From HS 2002"].isna(), "From HS 2002"] = 271000
            st.session_state.df_m = st.session_state.df_m.merge(aipnet, how="left", left_on="From HS 2002", right_on="hs_code_6d").drop(columns=["From HS 2022", "From HS 2002", "hs_code_6d"])
            st.session_state.modified_z_infl = False
            st.rerun()
            

if not st.session_state.modified_z_infl:
    with st.expander("**Sélectionner un pays à étudier**", expanded=True):

        selected_country = st.selectbox(
            "Pays à étudier",
            options=sorted(list(c_config.loc[c_config["zone"].isna(), "nom_pays"])),
            key="selected_country",
            on_change=lambda: st.session_state.update(modified_sel_country=True)
        )
        if st.session_state.modified_sel_country:
            selected_country_code = c_config.loc[c_config["nom_pays"] == selected_country, "country_code"].iloc[0]
            if st.button("🔍 **Appliquer les filtres**"):
                st.session_state.df_final = calc_var("imp", selected_country_code).merge(calc_var("exp", selected_country_code), how="outer").merge(st.session_state.df_m, how="outer")
                for col in st.session_state.df_final.columns:
                    if (col[0] == "v") | (col[0] == "q") | (col[0] == "p"):
                        st.session_state.df_final[col] = st.session_state.df_final[col].fillna(0)
                if c_config.loc[c_config["country_code"] == 251, "zone"].isna().all():
                    st.session_state.fr_ue_lab = "la France"
                else:
                    st.session_state.fr_ue_lab = "l'UE"
                st.session_state.df_final.rename(columns={"k":"Code HS6"}, inplace=True)
                st.session_state.df_final["Code HS6"] = st.session_state.df_final["Code HS6"].apply(conv_str)
                st.session_state.df_final = st.session_state.df_final.merge(labels, how="left")
                st.session_state.df_final["Code HS4"] = st.session_state.df_final["Code HS6"].apply(lambda x : x[:4])
                st.session_state.df_final = st.session_state.df_final.merge(labels.rename(columns={"Code HS6":"Code HS4", "Label HS6":"Label HS4"}), how="left")
                st.session_state.df_final.columns = ["Code HS6", "Importations du pays (en 1000$)", "Quantités importées du pays (en tonnes)", "HHi des importations du pays",
                    "Part du premier exportateur dans les importations du pays (en %)", "Part du deuxième exportateur dans les importations du pays (en %)", "Part du troisième exportateur dans les importations du pays (en %)", 
                    "Premier exportateur dans les importations du pays", "Deuxième exportateur dans les importations du pays", "Troisième exportateur dans les importations du pays", 
                    f"Part des exportations de {st.session_state.fr_ue_lab} dans les importations du pays (en %)", f"Part des importations du pays dans les exportations de {st.session_state.fr_ue_lab} (en %)", 
                    "Exportations du pays (en 1000$)", "Quantités exportées du pays (en tonnes)", "HHi des exportations du pays",
                    "Part du premier importateur dans les exportations du pays (en %)", "Part du deuxième importateur dans les exportations du pays (en %)", "Part du troisième importateur dans les exportations du pays (en %)", 
                    "Premier importateur dans les exportations du pays", "Deuxième importateur dans les exportations du pays", "Troisième importateur dans les exportations du pays",
                    f"Part des importations de {st.session_state.fr_ue_lab} dans les exportations du pays (en %)", f"Part des exportations du pays dans les importations de {st.session_state.fr_ue_lab} (en %)", 
                    "Valeur des flux échangés dans le monde (en 1000$)", "Quantités échangées dans le monde", "HHi des exportations mondiales",
                    "Part du premier exportateur mondial (en %)", "Part du deuxième exportateur mondial (en %)", "Part du troisième exportateur mondial (en %)",
                    "Premier exportateur mondial", "Deuxième exportateur mondial", "Troisième exportateur mondial",
                    f"Part des exportations de {st.session_state.fr_ue_lab} dans le monde (en %)",
                    "a supprimer", "a supprimer 2", "HHi des importations mondiales",
                    "Part du premier importateur mondial (en %)", "Part du deuxième importateur mondial (en %)", "Part du troisième importateur mondial (en %)",
                    "Premier importateur mondial", "Deuxième importateur mondial", "Troisième importateur mondial",
                    f"Part des importations de {st.session_state.fr_ue_lab} dans le monde (en %)",
                    "IGPC", "IGPC_rank",
                    "Label HS6", "Code HS4", "Label HS4"]
                l_cols = ["Code HS6", "Label HS6",
                    f"Part des exportations de {st.session_state.fr_ue_lab} dans les importations du pays (en %)", f"Part des importations du pays dans les exportations de {st.session_state.fr_ue_lab} (en %)", "HHi des importations du pays",
                    f"Part des importations de {st.session_state.fr_ue_lab} dans les exportations du pays (en %)", f"Part des exportations du pays dans les importations de {st.session_state.fr_ue_lab} (en %)", "HHi des exportations du pays",
                    "HHi des exportations mondiales", "HHi des importations mondiales",
                    "Importations du pays (en 1000$)", "Quantités importées du pays (en tonnes)",
                    "Exportations du pays (en 1000$)", "Quantités exportées du pays (en tonnes)",
                    "Valeur des flux échangés dans le monde (en 1000$)", "Quantités échangées dans le monde (en tonnes)",
                    f"Part des exportations de {st.session_state.fr_ue_lab} dans le monde (en %)", f"Part des importations de {st.session_state.fr_ue_lab} dans le monde (en %)",
                    "Code HS4", "Label HS4", "IGPC", "IGPC_rank",
                    "Part du premier exportateur dans les importations du pays (en %)", "Part du deuxième exportateur dans les importations du pays (en %)", "Part du troisième exportateur dans les importations du pays (en %)",          
                    "Part du premier importateur dans les exportations du pays (en %)", "Part du deuxième importateur dans les exportations du pays (en %)", "Part du troisième importateur dans les exportations du pays (en %)", 
                    "Part du premier exportateur mondial (en %)", "Part du deuxième exportateur mondial (en %)", "Part du troisième exportateur mondial (en %)",
                    "Part du premier importateur mondial (en %)", "Part du deuxième importateur mondial (en %)", "Part du troisième importateur mondial (en %)",
                    "Premier exportateur dans les importations du pays", "Deuxième exportateur dans les importations du pays", "Troisième exportateur dans les importations du pays", 
                    "Premier importateur dans les exportations du pays", "Deuxième importateur dans les exportations du pays", "Troisième importateur dans les exportations du pays", 
                    "Premier exportateur mondial", "Deuxième exportateur mondial", "Troisième exportateur mondial",
                    "Premier importateur mondial", "Deuxième importateur mondial", "Troisième importateur mondial"]
                st.session_state.df_final = st.session_state.df_final[l_cols]
                st.session_state.modified_sel_country = False
                st.rerun()


    if not st.session_state.modified_sel_country:
        df_final_mod = st.session_state.df_final.copy()
        st.subheader("Filtres")

        with st.expander("Sélectionner les sections ou catégories HS2 à afficher", expanded=True, key="container_labels_sections_tree"):
            selected_categories_index = sac.tree(labels_sections_tree, index = checked_sections, open_index=[], checkbox=True, return_index = True, key="labels_sections_tree", height=500)
            selected_categories = labels_sections.loc[selected_categories_index, "Catégorie"].values
            df_final_mod = df_final_mod[df_final_mod["Code HS6"].apply(lambda x : x[:2] in selected_categories)]
        
        approche_filter = st.radio(
            "Approche :",
            options=["Pas d'approche", "Par les importations", "Par les exportations"],
            index=0
        )
        filtres= {}
        if approche_filter == "Par les importations":
            type, type2 = "imp", "exp"
            type_produit = "_importations"
            filtres["approche_filter"] = approche_filter
        elif approche_filter == "Par les exportations":
            type, type2 = "exp", "imp"
            type_produit = "_exportations"
            filtres["approche_filter"] = approche_filter
        else :
            type_produit = ""
        
        type_filter = st.radio(
            "S'intéresser aux produits :",
            options=["Tous", "Tels que le pays est importateur net", "Tels que le pays est exportateur net"],
            index=0
        )
        if type_filter == "Tels que le pays est importateur net":
            df_final_mod = df_final_mod[(df_final_mod["Importations du pays"] >= df_final_mod["Exportations du pays"])&(df_final_mod["Importations du pays"] > 0)]
            filtres["type_filter"] = type_filter
        elif type_filter == "Tels que le pays est exportateur net":
            df_final_mod = df_final_mod[(df_final_mod["Exportations du pays"] >= df_final_mod["Importations du pays"])&(df_final_mod["Exportations du pays"] > 0)]
            filtres["type_filter"] = type_filter
        if approche_filter != "Pas d'approche" :
            l_cols_2 = ["Code HS6", "Label HS6",
                f"Part des {type2}ortations de {st.session_state.fr_ue_lab} dans les {type}ortations du pays (en %)", f"Part des {type}ortations du pays dans les {type2}ortations de {st.session_state.fr_ue_lab} (en %)", f"HHi des {type}ortations du pays",
                f"HHi des {type2}ortations mondiales", f"{maj(type)}ortations du pays (en 1000$)", f"Quantités {type}ortées du pays (en tonnes)", f"{maj(type2)}ortations du pays (en 1000$)", f"Quantités {type2}ortées du pays (en tonnes)",
                "Valeur des flux échangés dans le monde (en 1000$)", "Quantités échangées dans le monde (en tonnes)",
                f"Part des {type2}ortations de {st.session_state.fr_ue_lab} dans le monde (en %)",
                "Code HS4", "Label HS4", "IGPC", "IGPC_rank",
                f"Part du premier {type2}ortateur dans les {type}ortations du pays (en %)", f"Part du deuxième {type2}ortateur dans les {type}ortations du pays (en %)", f"Part du troisième {type2}ortateur dans les {type}ortations du pays (en %)",          
                f"Part du premier {type2}ortateur mondial (en %)", f"Part du deuxième {type2}ortateur mondial (en %)", f"Part du troisième {type2}ortateur mondial (en %)",
                f"Premier {type2}ortateur dans les {type}ortations du pays", f"Deuxième {type2}ortateur dans les {type}ortations du pays", f"Troisième {type2}ortateur dans les {type}ortations du pays", 
                f"Premier {type2}ortateur mondial", f"Deuxième {type2}ortateur mondial", f"Troisième {type2}ortateur mondial"]
            df_final_mod = df_final_mod[l_cols_2]
        
            filter_by_hhi_c = st.checkbox(f"Filtrer les produits selon l'indice HHi des {type}ortations du pays", key="filter_by_hhi_c", help=get_explication_filtre("hhi_c", type, type2))
            if filter_by_hhi_c:
                hhi_c = st.slider(f"HHi des {type}ortations supérieur à :", min_value=0.0, max_value=1.0, value=0.25, step=0.01, format="%.2f")
                df_final_mod = df_final_mod[df_final_mod[f"HHi des {type}ortations du pays"] >= hhi_c]
                filtres["hhi_c"] = hhi_c

            filter_by_hhi_m = st.checkbox(f"Filtrer les produits selon l'indice HHi mondial des {type2}ortations", key="filter_by_hhi_m", help=get_explication_filtre("hhi_M", type, type2))
            if filter_by_hhi_m:
                hhi_M = st.slider(f"HHi mondial des {type2}ortations supérieur à :", min_value=0.0, max_value=1.0, value=0.25, step=0.01, format="%.2f")
                df_final_mod = df_final_mod[df_final_mod[f"HHi des {type2}ortations mondiales"] >= hhi_M]
                filtres["hhi_M"] = hhi_M

            filter_by_p_fr_ue_in_c = st.checkbox(f"Filtrer les produits selon la part des {type2}ortations de {st.session_state.fr_ue_lab} dans les {type}ortations du pays", key="filter_by_p_fr_ue_in_c", help=get_explication_filtre("p_fr_ue_in_c", type, type2))
            if filter_by_p_fr_ue_in_c:
                p_fr_ue_in_c = st.slider(f"Part des {type2}ortations de {st.session_state.fr_ue_lab} dans les {type}ortations du pays supérieure à :", min_value=0, max_value=100, value=10, step=1, format="%d %%")
                df_final_mod = df_final_mod[df_final_mod[f"Part des {type2}ortations de {st.session_state.fr_ue_lab} dans les {type}ortations du pays (en %)"] >= p_fr_ue_in_c]
                filtres["p_fr_ue_in_c"] = f"{p_fr_ue_in_c}%"

            filter_by_p_c_in_fr_ue = st.checkbox(f"Filtrer les produits selon la part des {type}ortations du pays dans les {type2}ortations de {st.session_state.fr_ue_lab}", key="filter_by_p_c_in_fr_ue", help=get_explication_filtre("p_c_in_fr_ue", type, type2))
            if filter_by_p_c_in_fr_ue:
                p_c_in_fr_ue = st.slider(f"Part des {type}ortations du pays dans les {type2}ortations de {st.session_state.fr_ue_lab} inférieure à :", min_value=0, max_value=100, value=50, step=1, format="%d %%")
                df_final_mod = df_final_mod[df_final_mod[f"Part des {type}ortations du pays dans les {type2}ortations de {st.session_state.fr_ue_lab} (en %)"] <= p_c_in_fr_ue]
                filtres["p_c_in_fr_ue"] = f"{p_c_in_fr_ue}%"

            filter_by_igpc_rank = st.checkbox(f"Filtrer les produits selon le rang IGPC (score de centralité d'AIPNET)", key="filter_by_igpc_rank", help=get_explication_filtre("p_c_in_fr_ue", type, type2))
            if filter_by_igpc_rank:
                if type == "imp":
                    igpc_rank = st.slider(f"Rang IGPC supérieur à :", min_value=0, max_value=100, value=50, step=1)
                    df_final_mod = df_final_mod[(df_final_mod["IGPC_rank"] >= igpc_rank) | (df_final_mod[f"IGPC_rank"].isna())]
                    filtres["igpc_rank"] = str(igpc_rank)
                else :
                    igpc_rank = st.slider(f"Rang IGPC inférieur à :", min_value=0, max_value=100, value=50, step=1)
                    df_final_mod = df_final_mod[(df_final_mod["IGPC_rank"] <= igpc_rank) | (df_final_mod[f"IGPC_rank"].isna())]
                    filtres["igpc_rank"] = str(igpc_rank)

            log_values = np.linspace(-3, 6, 901)
            filter_by_v = st.checkbox(f"Filtrer les produits selon le montant des {type}ortations du pays :", key="filter_by_v")
            if filter_by_v:
                log_value_v = st.select_slider(f"Montant des {type}ortations (en 1000$) supérieur à :", options=log_values, value=0, format_func=lambda x: f"{10**x:.3f}")
                df_final_mod = df_final_mod[df_final_mod[f"{maj(type)}ortations du pays"] >= 10**log_value_v]
                filtres["v"] = f"{10**log_value_v * 1000}$"
            
        st.subheader("Résultats")
        st.write(f"Il y a {len(df_final_mod)} produits (HS6)")
        st.dataframe(df_final_mod)

        st.write("**Exportations :**")
        nom_fichier = st.text_input("Nom du fichier à télécharger (sans extension) :", value=f"dependances_inversees_{selected_country.replace(' ', '_')}{type_produit}", key=f"nom_fichier_{selected_country}{type_produit}")
        st.download_button(
            label="📥 Télécharger le tableau filtré en csv (sans les métadonnées et informations sur les filtres)",
            data=df_final_mod.to_csv(index=False, sep=";").encode("utf-8"),
            file_name=f"{nom_fichier}.csv",
            mime="text/csv"
        )

        if st.button("🔧 Préparer le téléchargement en excel"):

            zones_influence = {}
            for zone in c_config["zone"].unique():
                if not np.isnan(zone):
                    zones_influence[c_config.loc[c_config["country_code"] == zone, "nom_pays"].values[0]] = sorted(list(c_config.loc[c_config["zone"] == zone, "nom_pays"]))
            
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

                ws = writer.book.create_sheet("Informations")

                titre_font = Font(size=16, bold=True)
                bold = Font(bold=True)

                ws.merge_cells("A1:C1")
                ws["A1"] = "Analyse des dépendances inversées"
                ws["A1"].font = titre_font
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

                row = 3

                infos = [
                    ("Nombre de produits (HS6)", len(df_final_mod)),
                    ("Source", "Données BACI V" + version_baci),
                    ("Année", annee_baci),
                    ("Version HS6", version_hs6),
                    ("Pays étudié", selected_country),
                    ("Version AIPNET", version_aipnet),
                    ("Nomenclature utilisée pour AIPNET", nomenclature_aipnet)
                ]

                for cle, valeur in infos:
                    ws[f"A{row}"] = cle
                    ws[f"A{row}"].font = bold
                    ws[f"B{row}"] = valeur
                    row += 1

                row += 1
                if zones_influence:
                    ws[f"A{row}"] = "Zones d'influence"
                    ws[f"A{row}"].font = bold
                    ws[f"B{row}"] = "Pays"
                    ws[f"B{row}"].font = bold
                    row += 1

                    for zone, pays in zones_influence.items():
                        ws[f"A{row}"] = zone
                        ws[f"B{row}"] = ", ".join(pays)
                        row += 1

                    row += 1
                if filtres :
                    lab_filtres = {"approche_filter" : "Approche",
                            "type_filter" : "Produits",
                            "hhi_c" : f"HHi des {type}ortations supérieur à",
                            "hhi_M" : f"HHi mondial des {type2}ortations supérieur à",
                            "p_fr_ue_in_c" : f"Part des {type2}ortations de {st.session_state.fr_ue_lab} dans les {type}ortations du pays supérieure à",
                            "p_c_in_fr_ue" : f"Part des {type}ortations du pays dans les {type2}ortations de {st.session_state.fr_ue_lab} inférieure à",
                            "v" : f"Montant des {type}ortations du pays supérieur à"}
                    if type == "imp":
                        lab_filtres["igpc_rank"]  = "Rang IGPC supérieur à"
                    else:
                        lab_filtres["igpc_rank"]  = "Rang IGPC inférieur à"
                    ws[f"A{row}"] = "Filtres appliqués"
                    ws[f"A{row}"].font = bold
                    ws[f"B{row}"] = "Valeur"
                    ws[f"B{row}"].font = bold
                    row_filtre = row
                    row += 1

                    for filtre, seuil in filtres.items():
                        ws[f"A{row}"] = lab_filtres.get(filtre, "")
                        ws[f"B{row}"] = seuil
                        ws[f"C{row}"] = get_explication_filtre(filtre, type, type2)
                        if len(get_explication_filtre(filtre, type, type2)) > 0:
                            ws[f"C{row_filtre}"] = "Explications"
                            ws[f"C{row_filtre}"].font = bold
                        row += 1
                    row += 1
                
                vert = PatternFill(fill_type="solid", fgColor="C6EFCE")
                orange = PatternFill(fill_type="solid", fgColor="FFF2CC")
                rouge = PatternFill(fill_type="solid", fgColor="F4CCCC")

                ws[f"A{row}"] = "Sections douanières"
                ws[f"A{row}"].font = bold
                ws[f"B{row}"] = "Nombre de catégories HS2 sélectionnées"
                ws[f"B{row}"].font = bold
                ws[f"C{row}"] = "Catégories HS2 sélectionnées"
                ws[f"C{row}"].font = bold
                row += 1
                set_selected = set(selected_categories)
                for section in labels_sections[labels_sections["Niveau"] == "Section"].iterrows():
                    ws[f"A{row}"] = "Section " + section[1].Catégorie + " - " + section[1].Label 
                    l_hs2 = section[1].l_hs2
                    ws[f"B{row}"] = f"{len(set_selected&set(l_hs2))} / {len(l_hs2)}"
                    ws[f"C{row}"] = ", ".join(sorted(list(set_selected&set(l_hs2))))
                    if len(set_selected&set(l_hs2)) == len(l_hs2):
                        couleur = vert
                    elif len(set_selected&set(l_hs2)) > 0 :
                        couleur = orange
                    else :
                        couleur = rouge
       
                    ws[f"A{row}"].fill = couleur
                    ws[f"B{row}"].fill = couleur
                    ws[f"C{row}"].fill = couleur
                    row += 1

                ws.column_dimensions["A"].width = 50
                ws.column_dimensions["B"].width = 50
                ws.column_dimensions["C"].width = 50
                df_final_mod.to_excel(writer, sheet_name="Résultats", index=False)
            buffer.seek(0)

            st.download_button(
                label="📥 Télécharger le tableau filtré",
                data=buffer,
                file_name=f"{nom_fichier}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
